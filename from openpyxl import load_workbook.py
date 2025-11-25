from openpyxl import load_workbook


        # data_only=True로 해줘야 수식이 아닌 값으로 받아온다. 
load_wb = load_workbook("C:/icon_db/202511/entax_splitted_202511.xlsx", data_only=True)
        # 시트 이름으로 불러
        #         # 오기 
ws_list=[]
ws_list=load_wb.sheetnames
for a_ws in ws_list:
       print (a_ws)


load_ws_field = load_wb['field']

        # 셀 주소로 값 출력
print(load_ws_field['B2'].value)

        # 셀 좌표로 값 출력
print(load_ws_field.cell(3, 2).value)


        # 지정한 셀의 값 출력

get_cells = load_ws_field['B3' : 'B6']
for row in get_cells:
    for cell in row:
                print(cell.value)

        # 모든 행 단위로 출력

for row in load_ws_field.rows:
            print(row)

        # 모든 열 단위로 출력

for column in load_ws_field.columns:
            print(column)

        # 모든 행과 열 출력

all_values = []
for row in load_ws_field.rows:
            row_value = []
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
for row in all_values:
        print(row)
print(all_values)


        