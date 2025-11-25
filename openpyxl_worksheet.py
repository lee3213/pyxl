from openpyxl import load_workbook
from openpyxl import Workbook
import copy


        # data_only=True로 해줘야 수식이 아닌 값으로 받아온다. 
file_path="C:/icon_db/202511/entax_splitted_202511.xlsx"
sheet_name="customer_index"
find_string="자운"
try:
    load_wb = load_workbook(file_path, data_only=True)
        # 시트 이름으로 불러
        #         # 오기 
    ws_customer_index=load_wb[sheet_name]
except FileNotFoundError:
    print(f"ERror: The File '{file_path}' was not found.")
except KeyError:
    print(f"Error The Sheet '{sheet_name}' was not found in worksheet.")
ws_list=[]
ws_list=load_wb.sheetnames
# for a_ws in ws_list:
#     print (a_ws)
ws_trade=load_wb['sorted_trade']
ws_trade_detail=load_wb['sorted_detail']
trade_col=10 
# id, trade_type,dateof trade,trade_code,biz_of seller, biz of buyer, product,unit price, count, sum of transaction
trade_detail_col=11

print("columns",ws_customer_index.columns)
print ("dimensions:",ws_customer_index.dimensions)
print("min_column",ws_customer_index.min_column)
print("min_row",ws_customer_index.min_row )
print("max_column",ws_customer_index.max_column)
print("max_row",ws_customer_index.max_row )
trade_detail={}
detail_list=[]
a_detail_list=[]

x_trade_id=0
detail_idx={}
trade_detail_range_value=ws_trade_detail.iter_rows(min_col=1,max_col=trade_detail_col,min_row=9000,max_row=ws_trade_detail.max_row,values_only=True)
for y,a_detail in enumerate(trade_detail_range_value):
    
    if y < 5:
        print(str(a_detail), end=", ")
    else:
        break
        
for x,a_detail in enumerate(trade_detail_range_value):

    
    if x < 5 :
        print ("\nx=",x)
    
    tr_id=a_detail[0]
        
    if tr_id != x_trade_id:
            print("trid:",tr_id,x,x_trade_id)
            detail_list_copy=copy.deepcopy(detail_list)
            detail_idx[tr_id]=detail_list_copy
            x_trade_id=tr_id

        #trade_detail[a_detail[1]]=a_detail_list
    detail_list.add(tr_id)
        
    a_detail_list={}
    a_detail_list["dateof_trade"]=a_detail[2]
    a_detail_list["product_code"]=a_detail[3]
    a_detail_list["subsctiption"]=a_detail[4]
    a_detail_list["standard"]=a_detail[5]
    a_detail_list["unit"]=a_detail[6]
    a_detail_list["count"]=a_detail[7]
    a_detail_list["unit_price"]=a_detail[8]
    a_detail_list["sum"]=a_detail[9]
    a_detail_list["tax"]=a_detail[10]
    trade_detail[tr_id]=copy.deepcopy(a_detail_list)
    a_detail_list.clear()

for x,a_detail_id in enumerate(trade_detail):
    if x < 5:
        print(x,a_detail,trade_detail[a_detail_id],end="  ")
    print("\n")
for z,a_detail_idx in detail_idx:
    if x < 5:
        print(z,a_detail_idx)

from dataclasses import dataclass

@dataclass
class Product:
    weight:int = None
    price:int = None

apple = Product()
apple.weight = 10

shop={}
shop_list=[]
for customer_range in ws_customer_index.iter_cols(min_col=4,max_col=4,min_row=1,max_row=ws_customer_index.max_row,values_only=True):
    
        for i,string in enumerate(customer_range):
            if find_string in string :
                col_index=customer_range.index
                print ("loc of ",find_string,i)
                row_all=ws_customer_index.iter_rows(min_col=1,max_col=8, min_row=i+1,max_row=i+1,values_only=True)
                
                for x,cell in enumerate(row_all):
                    len_find=len(cell)
                    print("x=",x,len_find)
                    shop['trade_start']=cell[5]
                    shop['trade_end']=cell[6]
                    shop['shop_info_loc']=cell[7]
                    shop['shop_name']=cell[3]
                    shop_list.append(shop)
                    #for z,r_value in enumerate(cell):
                    #
                    #    print( z,r_value)

for a_shop in shop_list:
    for shop_key,shop_hint in a_shop.items():
        print(shop_key,shop_hint)
    trade_start_idx=a_shop['trade_start']
    trade_end_idx=a_shop['trade_start']
    for x,trade_range in enumerate(ws_trade.iter_rows(min_col=1,max_col=trade_col,min_row=trade_start_idx,max_row=trade_end_idx,values_only=True)):
        print("trade_idx=",x, "len=",len(trade_range))
        for y,a_trade in enumerate(trade_range):
            print ("y=",y,a_trade)
            a_detail_list=detail_list[a_trade[1]]

"""
            a_detail_list["dateof_trade"]=a_detail[2]
            a_detail_list["product_code"]=a_detail[3]
            a_detail_list["subsctiption"]=a_detail[4]
            a_detail_list["standard"]=a_detail[5]
            a_detail_list["unit"]=a_detail[6]
            a_detail_list["count"]=a_detail[7]
            a_detail_list["unit_price"]=a_detail[8]
            a_detail_list["sum"]=a_detail[9]
            a_detail_list["tax"]=a_detail[10]
"""